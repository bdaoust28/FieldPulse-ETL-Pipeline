import openpyxl,csv,re

# Step 3: input values into XLSX spreadsheet for upload
def input():
    xlsx = openpyxl.load_workbook('invoice_item_template.xlsx')
    sheet = xlsx.active
    items = csv.reader(open('swap.csv', newline='', encoding='utf8'), delimiter=',')
    skuCSV = csv.reader(open('sku.csv', newline='', encoding='utf8'), delimiter=',')
    skus,skuNum = {},{}     # second dict used to track SKU number per appliance category
    for item in skuCSV:
        skus[item[0]] = item[1] # {Refrigerators & Freezers: REFR}
        skuNum[item[1]] = 1000  # {REFR: 1000}
    # label columns
    col = {
        'name': 1,
        'sku': 2,
        'type': 3,
        'price': 5,
        'taxed': 7,
        'tags': 11
    }
    ignore = ['and', 'if', 'with', 'or', 'in', 'to', 'be', 'of']    # exclude these words from tags
    current_row = 3     # skips top 2 rows
    for row in items:
        try:
            sheet.cell(row = current_row, column = col['name']).value = f'{row[0]} - {row[1]}: {row[2]}'    # format of item

            # generate SKUs
            sheet.cell(row=current_row, column=col['sku']).value = skus[row[0]] + str(skuNum[skus[row[0]]]) # REFR1000
            skuNum[skus[row[0]]] += 1   # increment SKU number

            sheet.cell(row = current_row, column = col['type']).value = 'Service'   # must be changed if job is not a service
            sheet.cell(row = current_row, column = col['price']).value = float(row[4])
            sheet.cell(row = current_row, column = col['taxed']).value = 'No'       # resale tax applies when billing client for parts
            tags = re.split(r"[^\w]", ','.join(row[:3]))    # use regular expression to only include words up to third item in CSV row
            copy = tags.copy()      # copy tags to preserve tags in memory, necessary for traversal
            for tag in tags:
                if tag == '':
                    copy.remove(tag)        # remove tag if blank
                if tag.lower() in ignore:   # remove tag if in ignore list
                    try:
                        copy.remove(tag)
                    except(ValueError):
                        tag[0] = tag[0].upper() # make first char uppercase
                        copy.remove(tag)
            tags = copy     # replace old tags with cleaned up tags
            tags = ','.join(tags)
            sheet.cell(row = current_row, column = col['tags']).value = tags
            current_row += 1    # traverse rows
        except(IndexError):
            print(row)      # for debugging

    xlsx.save('fieldpulse-upload.xlsx')

if __name__ == '__main__':
    input()